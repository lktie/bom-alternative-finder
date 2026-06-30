#!/usr/bin/env python3
"""
excel_output.py — Generate BOM alternatives Excel output.

Column layout:
  Fixed (11 cols):
    Item Number | Description | Ref Des (BOM1) | Ref Des (BOM2) |
    Original Mfr 1 | Original MPN 1 | Orig 1 Characteristics |
    Original Mfr 2 | Original MPN 2 | Orig 2 Characteristics |
    Key Spec (Extracted)

  Per alternative (6 cols each, up to 10):
    Alt MPN | Alt Manufacturer | Datasheet URL | DigiKey URL | Qty Available | Remarks

Usage:
    from excel_output import write_bom_excel
    write_bom_excel(bom_rows, output_path)
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Constants ─────────────────────────────────────────────────────────────────

FIXED_HEADERS = [
    "Item Number",
    "Description",
    "Ref Des (BOM 1)",
    "Ref Des (BOM 2)",
    "Original Mfr 1",
    "Original MPN 1",
    "Orig 1 Characteristics",
    "Original Mfr 2",
    "Original MPN 2",
    "Orig 2 Characteristics",
    "Key Spec (Extracted)",
]

ALT_HEADERS = [
    "Alt MPN",
    "Alt Manufacturer",
    "Datasheet URL",
    "DigiKey URL",
    "Remarks",
]

MAX_ALTERNATIVES = 10

# ── Styles ────────────────────────────────────────────────────────────────────

HEADER_FILL    = PatternFill("solid", start_color="1F3864")   # dark navy
ALT_GRP_FILLS  = [
    PatternFill("solid", start_color="D9E1F2"),  # light blue  (odd alt groups)
    PatternFill("solid", start_color="E2EFDA"),  # light green (even alt groups)
]
HEADER_FONT    = Font(name="Arial", bold=True, color="FFFFFF", size=10)
BODY_FONT      = Font(name="Arial", size=10)
REMARK_FONT    = Font(name="Arial", size=10, italic=True, color="C00000")  # dark red
VERIFY_FILL    = PatternFill("solid", start_color="FFFF00")  # yellow — needs manual verification
WRAP           = Alignment(wrap_text=True, vertical="top")
THIN           = Side(style="thin", color="BFBFBF")
BORDER         = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _apply_header_style(cell, fill):
    cell.font      = HEADER_FONT
    cell.fill      = fill
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border    = BORDER


def _apply_body_style(cell, fill=None):
    cell.font      = BODY_FONT
    cell.alignment = WRAP
    cell.border    = BORDER
    if fill:
        cell.fill = fill


# ── Column width map ──────────────────────────────────────────────────────────

FIXED_COL_WIDTHS = [14, 40, 25, 25, 20, 22, 30, 20, 22, 30, 35]
ALT_COL_WIDTHS   = [22, 22, 35, 35, 40]


# ── Format key specs ──────────────────────────────────────────────────────────

def _format_specs(specs: dict) -> str:
    """Convert spec dict to readable multiline string."""
    lines = []
    for k, v in specs.items():
        if v and v != "-":
            lines.append(f"{k.replace('_', ' ').title()}: {v}")
    return "\n".join(lines)


# ── Main writer ───────────────────────────────────────────────────────────────

def write_bom_excel(bom_rows: list[dict], output_path: str):
    """
    Write BOM alternatives to Excel.

    Args:
        bom_rows: list of dicts, each representing one BOM line item:
            {
                "item_number":      str,
                "description":      str,
                "ref_des_bom1":     str,
                "ref_des_bom2":     str,
                "orig_mfr1":        str,
                "orig_mpn1":        str,
                "orig_chars1":      str,
                "orig_mfr2":        str,   # optional
                "orig_mpn2":        str,   # optional
                "orig_chars2":      str,   # optional
                "key_spec":         str,   # extracted spec string
                "alternatives": [
                    {
                        "mpn":               str,
                        "manufacturer":      str,
                        "datasheet_url":     str,
                        "product_url":       str,
                        "quantity_available": int,
                        "remarks":           str,   # empty = all match
                    },
                    ...
                ]
            }
        output_path: where to save the .xlsx file
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "BOM Alternatives"
    ws.freeze_panes = "A2"  # freeze header row

    # ── Build header row ──────────────────────────────────────────────────────
    header_row = FIXED_HEADERS[:]
    for i in range(1, MAX_ALTERNATIVES + 1):
        for h in ALT_HEADERS:
            header_row.append(f"Alt {i} — {h}")

    for col_idx, header in enumerate(header_row, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        # Determine fill: fixed headers get navy, alt groups alternate
        if col_idx <= len(FIXED_HEADERS):
            fill = HEADER_FILL
        else:
            alt_idx = (col_idx - len(FIXED_HEADERS) - 1) // len(ALT_HEADERS)
            fill = PatternFill("solid",
                               start_color="2E5C8A" if alt_idx % 2 == 0 else "27674A")
        _apply_header_style(cell, fill)

    ws.row_dimensions[1].height = 40

    # ── Set column widths ─────────────────────────────────────────────────────
    for i, width in enumerate(FIXED_COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    for alt_i in range(MAX_ALTERNATIVES):
        base = len(FIXED_HEADERS) + alt_i * len(ALT_HEADERS)
        for j, width in enumerate(ALT_COL_WIDTHS):
            ws.column_dimensions[get_column_letter(base + j + 1)].width = width

    # ── Write data rows ───────────────────────────────────────────────────────
    for row_idx, item in enumerate(bom_rows, 2):
        alts = item.get("alternatives", [])

        # Determine row fill: alternate light grey / white
        row_fill = PatternFill("solid", start_color="F2F2F2") if row_idx % 2 == 0 else None

        fixed_values = [
            item.get("item_number", ""),
            item.get("description", ""),
            item.get("ref_des_bom1", ""),
            item.get("ref_des_bom2", ""),
            item.get("orig_mfr1", ""),
            item.get("orig_mpn1", ""),
            item.get("orig_chars1", ""),
            item.get("orig_mfr2", ""),
            item.get("orig_mpn2", ""),
            item.get("orig_chars2", ""),
            item.get("key_spec", ""),
        ]

        for col_idx, value in enumerate(fixed_values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            _apply_body_style(cell, row_fill)

        # Write alternative columns
        for alt_i, alt in enumerate(alts[:MAX_ALTERNATIVES]):
            base_col = len(FIXED_HEADERS) + alt_i * len(ALT_HEADERS) + 1
            alt_fill = ALT_GRP_FILLS[alt_i % 2]

            remarks = alt.get("remarks", "")

            datasheet_verify = alt.get("datasheet_verify", False)
            alt_values = [
                alt.get("mpn", ""),
                alt.get("manufacturer", ""),
                alt.get("datasheet_url", ""),
                alt.get("product_url", ""),
                remarks,
            ]

            for j, value in enumerate(alt_values):
                cell = ws.cell(row=row_idx, column=base_col + j, value=value)
                if j == 2 and datasheet_verify:
                    # Datasheet URL needs manual verification — highlight yellow
                    cell.font      = BODY_FONT
                    cell.fill      = VERIFY_FILL
                    cell.alignment = WRAP
                    cell.border    = BORDER
                elif j == 4 and remarks:
                    # Remarks cell — dark red italic
                    cell.font      = REMARK_FONT
                    cell.fill      = alt_fill
                    cell.alignment = WRAP
                    cell.border    = BORDER
                else:
                    _apply_body_style(cell, alt_fill)

        # Fill remaining alt columns with empty styled cells
        for alt_i in range(len(alts), MAX_ALTERNATIVES):
            base_col = len(FIXED_HEADERS) + alt_i * len(ALT_HEADERS) + 1
            alt_fill = ALT_GRP_FILLS[alt_i % 2]
            for j in range(len(ALT_HEADERS)):
                cell = ws.cell(row=row_idx, column=base_col + j, value="")
                _apply_body_style(cell, alt_fill)

        ws.row_dimensions[row_idx].height = 60

    wb.save(output_path)
    print(f"[excel] Saved → {output_path}")


# ── Standalone test ───────────────────────────────────────────────────────────

if __name__ == "__main__":
    import os
    from spec_extractor import extract_specs
    from alternatives_fetcher import get_token, fetch_alternatives
    from remarks_generator import generate_remarks

    CLIENT_ID     = os.getenv("DIGIKEY_CLIENT_ID")
    CLIENT_SECRET = os.getenv("DIGIKEY_CLIENT_SECRET")

    # Minimal test: two MPNs, simulating a BOM ingestion stub
    # In the real tool this will come from the BoM parser
    TEST_BOM = [
        {
            "item_number":  "120-00052",
            "description":  "2.2µF ±10% 10V Ceramic Capacitor X5R 0201 (0603 Metric)",
            "ref_des_bom1": "C18, C19, C21",
            "ref_des_bom2": "",
            "orig_mfr1":    "Murata Manufacturing",
            "orig_mpn1":    "GRM033R61A225KE47D",
            "orig_chars1":  "Capacitor, 2.2uF, ±10%, 10V, X5R, 0201",
            "orig_mfr2":    "",
            "orig_mpn2":    "",
            "orig_chars2":  "",
        },
        {
            "item_number":  "180-00118",
            "description":  "0 Ohms Jumper 0.05W, 1/20W Chip Resistor 0201 (0603 Metric) Automotive AEC-Q200 Thick Film",
            "ref_des_bom1": "R2, R17, R94, R117, R118, R150, R151",
            "ref_des_bom2": "",
            "orig_mfr1":    "Panasonic",
            "orig_mpn1":    "ERJ-1GN0R00C",
            "orig_chars1":  "Resistor, 0 Ohms, Jumper, 0201",
            "orig_mfr2":    "",
            "orig_mpn2":    "",
            "orig_chars2":  "",
        },
    ]

    token = get_token()
    print("[token] OK\n")

    bom_rows = []
    for item in TEST_BOM:
        mpn = item["orig_mpn1"]
        print(f"Processing {mpn}...")
        results = fetch_alternatives(token, mpn)

        if "error" in results:
            print(f"  ERROR: {results['error']}")
            item["key_spec"]     = ""
            item["alternatives"] = []
        else:
            original  = results["original"]
            comp_type = original["component_type"]
            item["key_spec"] = _format_specs(original.get("specs", {}))

            alts = []
            for alt in results["alternatives"]:
                note = generate_remarks(original, alt, comp_type)
                alts.append({
                    "mpn":                alt["mpn"],
                    "manufacturer":       alt["manufacturer"],
                    "datasheet_url":      alt.get("datasheet_url", ""),
                    "datasheet_verify":   alt.get("datasheet_verify", False),
                    "product_url":        alt.get("product_url", ""),
                    "remarks":            "" if note == "All key specs match original." else note,
                })
            item["alternatives"] = alts

        bom_rows.append(item)

    output_path = os.path.join(os.path.dirname(__file__), "bom_alternatives_output.xlsx")
    write_bom_excel(bom_rows, output_path)
    print(f"\nDone. Open: {output_path}")
