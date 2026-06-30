#!/usr/bin/env python3
"""
bom_ingestion.py — Main entry point for BOM alternatives tool.

Usage:
    export DIGIKEY_CLIENT_ID="..."
    export DIGIKEY_CLIENT_SECRET="..."
    python bom_ingestion.py bom1.xlsx bom2.xlsx

Flow:
    1. Read all input BOM files
    2. Filter to 120- and 180- item numbers
    3. Merge and deduplicate by item number
    4. Run each MPN through alternatives pipeline
    5. Write output Excel with dynamic manufacturer brand columns
"""

import os
import sys
import re
import json
import time
from pathlib import Path
from collections import OrderedDict

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from alternatives_fetcher import get_token, fetch_alternatives
from remarks_generator import generate_remarks
from spec_extractor import extract_specs

# ── Config ────────────────────────────────────────────────────────────────────

ITEM_PREFIXES    = ("120-", "180-")
MAX_ALTERNATIVES = 10
CHECKPOINT_FILE  = ".bom_checkpoint.json"
MAX_RETRIES      = 3
RETRY_DELAY      = 5  # seconds

# Column positions in input BOM (0-indexed)
COL_ITEM_NUM  = 0   # A - Item Number
COL_DESC      = 1   # B - Description
COL_REF_DES   = 2   # C - Ref Des (first BoM, may vary)
COL_ORIG_MFR1 = 4   # E - Original Mfr 1
COL_ORIG_MPN1 = 5   # F - Original MPN 1
COL_ORIG_CH1  = 6   # G - Orig 1 Characteristics
COL_ORIG_MFR2 = 7   # H - Original Mfr 2 (optional)
COL_ORIG_MPN2 = 8   # I - Original MPN 2 (optional)
COL_ORIG_CH2  = 9   # J - Orig 2 Characteristics (optional)

# ── Styles ────────────────────────────────────────────────────────────────────

HEADER_FILL   = PatternFill("solid", start_color="1F3864")
MFR_FILLS     = [
    PatternFill("solid", start_color="D9E1F2"),
    PatternFill("solid", start_color="E2EFDA"),
]
HEADER_FONT   = Font(name="Arial", bold=True, color="FFFFFF", size=10)
BODY_FONT     = Font(name="Arial", size=10)
REMARK_FONT   = Font(name="Arial", size=10, italic=True, color="C00000")
VERIFY_FILL   = PatternFill("solid", start_color="FFFF00")
THIN          = Side(style="thin", color="BFBFBF")
BORDER        = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP          = Alignment(wrap_text=True, vertical="top")
ROW_FILLS     = [PatternFill("solid", start_color="F2F2F2"), None]


def _header_style(cell, fill):
    cell.font      = HEADER_FONT
    cell.fill      = fill
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border    = BORDER


def _body_style(cell, fill=None, font=None):
    cell.font      = font or BODY_FONT
    cell.alignment = WRAP
    cell.border    = BORDER
    if fill:
        cell.fill = fill


# ── Step 1: Read and merge BOM files ─────────────────────────────────────────

def read_bom_file(filepath: str) -> tuple[str, pd.DataFrame]:
    """Read a BOM Excel file. Returns (bom_number, dataframe)."""
    bom_number = Path(filepath).stem  # filename without extension
    df = pd.read_excel(filepath, header=0, dtype=str)
    df = df.fillna("")
    return bom_number, df


def merge_bom_files(filepaths: list[str]) -> dict:
    """
    Merge multiple BOM files into one dict keyed by item number.
    Each entry contains all fixed fields + ref des per BOM file.

    Returns:
        {
            item_number: {
                "item_number":  str,
                "description":  str,
                "orig_mfr1":    str,
                "orig_mpn1":    str,
                "orig_chars1":  str,
                "orig_mfr2":    str,
                "orig_mpn2":    str,
                "orig_chars2":  str,
                "ref_des":      { bom_number: ref_des_str },
            }
        }
    """
    merged = OrderedDict()

    for filepath in filepaths:
        bom_number, df = read_bom_file(filepath)
        print(f"[bom] Read {filepath} → {len(df)} rows, BOM: {bom_number}")

        for _, row in df.iterrows():
            vals = list(row)

            item_num = str(vals[COL_ITEM_NUM]).strip() if len(vals) > COL_ITEM_NUM else ""
            if not item_num or not any(item_num.startswith(p) for p in ITEM_PREFIXES):
                continue

            ref_des   = str(vals[COL_REF_DES]).strip()   if len(vals) > COL_REF_DES   else ""
            desc      = str(vals[COL_DESC]).strip()       if len(vals) > COL_DESC      else ""
            orig_mfr1 = str(vals[COL_ORIG_MFR1]).strip() if len(vals) > COL_ORIG_MFR1 else ""
            orig_mpn1 = str(vals[COL_ORIG_MPN1]).strip() if len(vals) > COL_ORIG_MPN1 else ""
            orig_ch1  = str(vals[COL_ORIG_CH1]).strip()  if len(vals) > COL_ORIG_CH1  else ""
            orig_mfr2 = str(vals[COL_ORIG_MFR2]).strip() if len(vals) > COL_ORIG_MFR2 else ""
            orig_mpn2 = str(vals[COL_ORIG_MPN2]).strip() if len(vals) > COL_ORIG_MPN2 else ""
            orig_ch2  = str(vals[COL_ORIG_CH2]).strip()  if len(vals) > COL_ORIG_CH2  else ""

            if item_num not in merged:
                merged[item_num] = {
                    "item_number": item_num,
                    "description": desc,
                    "orig_mfr1":   orig_mfr1,
                    "orig_mpn1":   orig_mpn1,
                    "orig_chars1": orig_ch1,
                    "orig_mfr2":   orig_mfr2,
                    "orig_mpn2":   orig_mpn2,
                    "orig_chars2": orig_ch2,
                    "ref_des":     {},
                }

            # Merge ref des for this BOM file
            if ref_des:
                existing = merged[item_num]["ref_des"].get(bom_number, "")
                if existing and ref_des not in existing:
                    merged[item_num]["ref_des"][bom_number] = existing + ", " + ref_des
                else:
                    merged[item_num]["ref_des"][bom_number] = ref_des

    print(f"[bom] {len(merged)} unique items after merge + dedup")
    return merged


# ── Step 2: Run pipeline ──────────────────────────────────────────────────────

def _format_key_spec(specs: dict) -> str:
    lines = []
    for k, v in specs.items():
        if v and v != "-":
            lines.append(f"{k.replace('_',' ').title()}: {v}")
    return "\n".join(lines)



# ── Checkpoint save/resume ────────────────────────────────────────────────────

def save_checkpoint(results: list[dict], processed_items: set):
    """Save progress to disk so a crash doesn't lose completed work."""
    data = {
        "results":         results,
        "processed_items": list(processed_items),
    }
    with open(CHECKPOINT_FILE, "w") as f:
        json.dump(data, f)


def load_checkpoint() -> tuple[list[dict], set]:
    """Load prior progress if checkpoint exists. Returns (results, processed_items)."""
    if not os.path.exists(CHECKPOINT_FILE):
        return [], set()
    try:
        with open(CHECKPOINT_FILE, "r") as f:
            data = json.load(f)
        results = data.get("results", [])
        processed = set(data.get("processed_items", []))
        print(f"[checkpoint] Resuming — {len(processed)} items already processed")
        return results, processed
    except Exception as e:
        print(f"[checkpoint] Could not load checkpoint: {e}")
        return [], set()


def clear_checkpoint():
    """Remove checkpoint file after successful completion."""
    if os.path.exists(CHECKPOINT_FILE):
        os.remove(CHECKPOINT_FILE)


def _fetch_with_retry(token, mpn: str, max_retries: int = MAX_RETRIES) -> dict:
    """Wrap fetch_alternatives with retry on network errors."""
    for attempt in range(1, max_retries + 1):
        try:
            return fetch_alternatives(token, mpn)
        except Exception as e:
            print(f"  [retry {attempt}/{max_retries}] {mpn} failed: {e}")
            if attempt < max_retries:
                time.sleep(RETRY_DELAY)
            else:
                return {"error": f"Failed after {max_retries} attempts: {e}"}

def process_items(token, merged_items: dict) -> list[dict]:
    """
    Run each item through the alternatives pipeline.
    Saves checkpoint after each item — resumes automatically if interrupted.
    Returns list of result dicts ready for Excel output.
    """
    results, processed = load_checkpoint()
    results_by_item = {r["item_number"]: r for r in results}

    total = len(merged_items)
    done  = len(processed)

    for item_num, item in merged_items.items():
        if item_num in processed:
            continue

        done += 1
        mpn = item["orig_mpn1"] or item["orig_mpn2"]
        if not mpn:
            print(f"[skip] {item_num} — no MPN")
            processed.add(item_num)
            continue

        print(f"\n[process {done}/{total}] {item_num} — {mpn}")

        pipeline_result = _fetch_with_retry(token, mpn)

        if "error" in pipeline_result:
            print(f"  [error] {pipeline_result['error']}")
            item["key_spec"]     = ""
            item["alternatives"] = []
            # Do NOT mark as processed — will retry on next run
            results_by_item[item_num] = item
            save_checkpoint(list(results_by_item.values()), processed)
            continue
        else:
            original  = pipeline_result["original"]
            comp_type = original["component_type"]
            item["key_spec"] = _format_key_spec(original.get("specs", {}))

            alts = []
            for alt in pipeline_result["alternatives"]:
                note = generate_remarks(original, alt, comp_type)
                alts.append({
                    "mpn":              alt["mpn"],
                    "manufacturer":     alt["manufacturer"],
                    "datasheet_url":    alt.get("datasheet_url", ""),
                    "datasheet_verify": alt.get("datasheet_verify", False),
                    "remarks":          "" if note == "All key specs match original." else note,
                })
            item["alternatives"] = alts

        results_by_item[item_num] = item
        processed.add(item_num)

        # Save checkpoint after every item
        save_checkpoint(list(results_by_item.values()), processed)

    return list(results_by_item.values())


# ── Step 3: Write output Excel ────────────────────────────────────────────────

def write_output(results: list[dict], bom_numbers: list[str], output_path: str):
    """Write results to Excel with dynamic manufacturer brand columns."""

    # Pass 1: collect all manufacturers in order of first appearance
    mfr_order = []
    seen_mfrs = set()
    for item in results:
        for alt in item.get("alternatives", []):
            mfr = alt["manufacturer"]
            if mfr not in seen_mfrs:
                seen_mfrs.add(mfr)
                mfr_order.append(mfr)

    print(f"\n[excel] Manufacturers: {mfr_order}")

    # Build fixed headers
    fixed_headers = ["Item Number", "Description"]
    for bom_num in bom_numbers:
        fixed_headers.append(f"Ref Des ({bom_num})")
    fixed_headers += [
        "Original Mfr 1", "Original MPN 1", "Orig 1 Characteristics",
        "Original Mfr 2", "Original MPN 2", "Orig 2 Characteristics",
        "Key Spec (Extracted)",
    ]

    # Manufacturer sub-headers (3 cols each)
    mfr_subheaders = ["MPN", "Datasheet URL", "Manufacturer"]

    wb = Workbook()
    ws = wb.active
    ws.title = "BOM Alternatives"
    ws.freeze_panes = "A3"  # freeze first 2 header rows

    # ── Row 1: main headers ───────────────────────────────────────────────────
    for col_idx, header in enumerate(fixed_headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        _header_style(cell, HEADER_FILL)

    for mfr_idx, mfr in enumerate(mfr_order):
        base_col = len(fixed_headers) + mfr_idx * 3 + 1
        fill = PatternFill("solid", start_color="2E5C8A" if mfr_idx % 2 == 0 else "27674A")
        # Merge 3 cells for manufacturer name
        ws.merge_cells(
            start_row=1, start_column=base_col,
            end_row=1,   end_column=base_col + 2
        )
        cell = ws.cell(row=1, column=base_col, value=mfr)
        _header_style(cell, fill)

    # ── Row 2: sub-headers for manufacturer columns ───────────────────────────
    # Fill fixed cols with empty styled cells
    for col_idx in range(1, len(fixed_headers) + 1):
        cell = ws.cell(row=2, column=col_idx, value="")
        _header_style(cell, HEADER_FILL)

    for mfr_idx, mfr in enumerate(mfr_order):
        base_col = len(fixed_headers) + mfr_idx * 3 + 1
        fill = PatternFill("solid", start_color="2E5C8A" if mfr_idx % 2 == 0 else "27674A")
        for j, sub in enumerate(mfr_subheaders):
            cell = ws.cell(row=2, column=base_col + j, value=sub)
            _header_style(cell, fill)

    ws.row_dimensions[1].height = 35
    ws.row_dimensions[2].height = 25

    # ── Column widths ─────────────────────────────────────────────────────────
    fixed_widths = [14, 40] + [20] * len(bom_numbers) + [20, 22, 30, 20, 22, 30, 35]
    for i, w in enumerate(fixed_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    for mfr_idx in range(len(mfr_order)):
        base_col = len(fixed_headers) + mfr_idx * 3 + 1
        for j, w in enumerate([22, 35, 22]):
            ws.column_dimensions[get_column_letter(base_col + j)].width = w

    # ── Data rows ─────────────────────────────────────────────────────────────
    # Build mfr lookup per item: { mfr_name: alt_dict }
    for row_idx, item in enumerate(results, 3):
        row_fill = ROW_FILLS[row_idx % 2]

        # Fixed columns
        fixed_values = [item.get("item_number", ""), item.get("description", "")]
        for bom_num in bom_numbers:
            fixed_values.append(item.get("ref_des", {}).get(bom_num, ""))
        fixed_values += [
            item.get("orig_mfr1", ""),
            item.get("orig_mpn1", ""),
            item.get("orig_chars1", ""),
            item.get("orig_mfr2", ""),
            item.get("orig_mpn2", ""),
            item.get("orig_chars2", ""),
            item.get("key_spec", ""),
        ]

        for col_idx, value in enumerate(fixed_values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            _body_style(cell, row_fill)

        # Build mfr → alt mapping for this item
        mfr_alt_map = {}
        for alt in item.get("alternatives", []):
            mfr = alt["manufacturer"]
            if mfr not in mfr_alt_map:
                mfr_alt_map[mfr] = alt

        # Write manufacturer columns
        for mfr_idx, mfr in enumerate(mfr_order):
            base_col = len(fixed_headers) + mfr_idx * 3 + 1
            alt_fill = MFR_FILLS[mfr_idx % 2]
            alt      = mfr_alt_map.get(mfr)

            if alt:
                mpn_cell = ws.cell(row=row_idx, column=base_col,     value=alt["mpn"])
                url_cell = ws.cell(row=row_idx, column=base_col + 1, value=alt["datasheet_url"])
                mfr_cell = ws.cell(row=row_idx, column=base_col + 2, value=alt["manufacturer"])

                _body_style(mpn_cell, alt_fill)
                _body_style(mfr_cell, alt_fill)

                # Datasheet URL — yellow if needs verification
                if alt.get("datasheet_verify"):
                    url_cell.font      = BODY_FONT
                    url_cell.fill      = VERIFY_FILL
                    url_cell.alignment = WRAP
                    url_cell.border    = BORDER
                else:
                    _body_style(url_cell, alt_fill)

                # Remarks as cell comment on MPN cell if non-empty
                if alt.get("remarks"):
                    from openpyxl.comments import Comment
                    comment = Comment(alt["remarks"], "Tool")
                    mpn_cell.comment = comment
            else:
                for j in range(3):
                    cell = ws.cell(row=row_idx, column=base_col + j, value="")
                    _body_style(cell, alt_fill)

        ws.row_dimensions[row_idx].height = 50

    wb.save(output_path)
    print(f"[excel] Saved → {output_path}")


# ── Main ──────────────────────────────────────────────────────────────────────

def main(input_files: list[str], output_path: str):
    if not input_files:
        print("Usage: python bom_ingestion.py bom1.xlsx [bom2.xlsx ...] [--output output.xlsx]")
        return

    # Step 1: merge
    merged = merge_bom_files(input_files)
    bom_numbers = [Path(f).stem for f in input_files]

    # Step 2: pipeline
    client_id     = os.getenv("DIGIKEY_CLIENT_ID")
    client_secret = os.getenv("DIGIKEY_CLIENT_SECRET")
    if not client_id or not client_secret:
        print("ERROR: set DIGIKEY_CLIENT_ID and DIGIKEY_CLIENT_SECRET")
        return

    from alternatives_fetcher import refresh_token
    token = refresh_token()
    print(f"[token] OK — processing {len(merged)} items\n")

    results = process_items(token, merged)

    # Step 3: write
    failed_items = [r["item_number"] for r in results if not r.get("alternatives") and not r.get("key_spec")]

    write_output(results, bom_numbers, output_path)

    if failed_items:
        print(f"\n[warn] {len(failed_items)} items failed: {failed_items}")
        print("[checkpoint] Kept — re-run the same command to retry failed items only")
    else:
        clear_checkpoint()
        print("[checkpoint] Cleared — run complete, no failures")


if __name__ == "__main__":
    args = sys.argv[1:]

    # Parse --output flag
    output_path = "bom_alternatives_output.xlsx"
    if "--output" in args:
        idx = args.index("--output")
        output_path = args[idx + 1]
        args = args[:idx] + args[idx + 2:]

    main(args, output_path)
